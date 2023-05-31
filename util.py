import openpyxl
import os
from collibra import *
def submit_prompt(prompt):
    insert_prompt_into_excel(prompt)
    return 'Prompt Saved Successfully'

def update_answer(prompt):
    update_answer_into_excel(prompt)
    return 'Answer Saved Successfully'

def insert_prompt_into_excel(prompt):
    # Open the Excel file
    wb = openpyxl.load_workbook('example.xlsx')

    # Select the first worksheet
    ws = wb.active

    prompt_cell = ws.cell(row=ws.max_row, column=1)
    prompt_cell.value = prompt

    # Save the changes to the Excel file
    wb.save('example.xlsx')
    wb.close()

def update_answer_into_excel(prompt):
    # Open the Excel file
    wb = openpyxl.load_workbook('C:/Users/shubham.prateek/Downloads/example.xlsx')

    # Select the first worksheet
    ws = wb.active

    prompt_cell = ws.cell(row=ws.max_row, column=2)
    prompt_cell.value = prompt

    # Save the changes to the Excel file
    wb.save('C:/Users/shubham.prateek/Downloads/example.xlsx')
    wb.close()

def refresh_output():
    file_path = r"C:\Users\shubham.prateek\Downloads\example.xlsx"

    # Check if the file exists
    if not os.path.isfile(file_path):
        return "Excel file not found"

    try:
        # Open the Excel file
        wb = openpyxl.load_workbook(file_path)

        # Select the first worksheet
        ws = wb.active

        # Find the value in the most recent row of the "Output" column
        row_num = ws.max_row
        output = ws.cell(row=row_num, column=2).value

        # Close the workbook
        wb.close()

        return output

    except FileNotFoundError:
        return "Excel file not found"

def Patch():
    attributeTypeId = {
        "Description": "00000000-0000-0000-0000-000000003114"
    }
    domainId = "fe3a443d-cb64-414c-b022-18d03436bcdb"

    # Read the outputs
    outs = pd.read_excel(r"C:\Users\shubham.prateek\Downloads\example.xlsx")
    #print(outs)
    # generate input structure
    df = pd.DataFrame(["t_product"], columns=['Column Name'])
    # Stores the assets UUID
    for i in range(df.shape[0]):
        df.loc[i, "Column Uuid"] = searchAssetUuid(df.loc[i, "Column Name"], domainId)

    if (outs.loc[0, "input"].find("technical") >= 0):
        print('hi')
        df["Tag"] = [outs.loc[0, "output"]]
        print(outs.loc[0, "output"])
        # Patch the df info to the respective assets in DGC
        word_list = df.iloc[0, 2].split(" ")
        print(word_list)
        for i in range(df.shape[0]):
            assetName = df.loc[0, "Column Name"]
            print(assetName)
            res = addTag(df.loc[0, "Column Uuid"], word_list)
            print(f"{assetName} Tag Attribute added: {res}")

    elif (outs.loc[0, "input"].find("business") >= 0):
        df["Description"] = outs.loc[0, "output"]
        # Patch the df info to the respective assets in DGC
        for i in range(df.shape[0]):
            assetName = df.loc[i, "Column Name"]
            for k, v in attributeTypeId.items():
                attributeId = searchAttributeUuid(df.loc[i, "Column Uuid"], k)
                if (k != "Description"):
                    val = bool(df.loc[i, k])
                else:
                    val = df.loc[i, k]

                if (attributeId != None):
                    res = setAttribute(attributeId, val)
                    print(f"{assetName} {k} Attribute set: {res}")
                else:
                    res = addAttribute(df.loc[i, "Column Uuid"], v, val)
                    print(f"{assetName} {k} Attribute added: {res}")
    print(df["Column Uuid"].tolist())
    print(type(df["Column Uuid"]))
    runApprovalWF(df["Column Uuid"].tolist())
    return attributeTypeId